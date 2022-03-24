from openpyxl import load_workbook as lwb
from openpyxl import Workbook
from os import listdir as ld

def add_row(wb,kad,ls,row):
    wb.cell(row,4).value = kad
    wb.cell(row,2).value = ls

list_files = ld()
print("lis")

for fl in list_files:
    if (".xlsx" in fl) and ("~" not in fl):
        print("fl = ", fl)
        wb_source = lwb(fl)
        ws_source = wb_source.active

        wb_result = Workbook()
        ws_result = wb_result.active

        i=1
        j=1

        ws_result.cell(1,1).value="Договор"
        ws_result.cell(1,2).value="Лицевой счет"
        ws_result.cell(1,3).value="ФИО"
        ws_result.cell(1,4).value="Адрес"

        max = ws_source.max_row
        print("max = ",max)

        while i <= max:
            i+=1
            #print(" i ", i)
            if (ws_source.cell(i,1).value != ws_source.cell(i-1,1).value):
                j+=1
                print("j = ",j)
                add_row(ws_result,ws_source.cell(i,6).value, ws_source.cell(i,7).value,j)
        file_name_result = fl.replace(".xlsx","_one.xlsx")

        wb_result.save(file_name_result)
print("the end")    